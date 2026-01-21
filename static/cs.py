import streamlit as st
from docx import Document
import os
import re
import json
import pandas as pd
from io import BytesIO
import tempfile
import zipfile
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# 把excel解析为特定格式json
def excel_to_json(excel_file, text_data_start, tables_data_mapping, sheetname=None):
    """
    将Excel文件转换为指定格式的JSON格式
    
    参数:
    excel_file: str - Excel文件路径
    text_data_start: str - text_data的左上角单元格，如"A1"
    tables_data_mapping: dict - tables_data的表名及其对应的左上角单元格，如{"家庭成员": "A10"}
    sheetname: str - 可选，指定要处理的sheet名称，如果不提供则处理所有sheet
    """
    # 读取Excel文件
    workbook = openpyxl.load_workbook(excel_file)
    
    result = []
    
    # 确定要处理的sheet列表
    sheet_names = [sheetname] if sheetname and sheetname in workbook.sheetnames else workbook.sheetnames
    
    # 遍历要处理的sheet
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        
        # 解析text_data起始单元格
        text_col = column_index_from_string(text_data_start[0])
        text_row = int(text_data_start[1:])
        
        if tables_data_mapping:
            # 有表格数据的情况：第一行为键，第二行为值，每个sheet生成一个文档
            # 处理text_data
            text_data = {}
            
            # text_data的键行和值行
            key_row = text_row
            value_row = text_row + 1
            current_col = text_col
            
            # 遍历所有列，直到遇到空单元格
            while True:
                key_cell = sheet.cell(row=key_row, column=current_col)
                value_cell = sheet.cell(row=value_row, column=current_col)
                
                # 检查是否为空单元格
                if key_cell.value is None:
                    break
                
                # 添加到text_data，键用{}包起来
                text_data[f"{{{key_cell.value}}}"] = value_cell.value
                current_col += 1
            
            # 处理tables_data
            tables_data = {}
            
            for table_name, table_start in tables_data_mapping.items():
                # 解析表格起始单元格
                table_col = column_index_from_string(table_start[0])
                table_row = int(table_start[1:])
                
                # 读取表头（起始行）
                headers = []
                header_row = table_row
                current_col = table_col
                
                while True:
                    header_cell = sheet.cell(row=header_row, column=current_col)
                    if header_cell.value is None:
                        break
                    headers.append(header_cell.value)
                    current_col += 1
                
                # 读取数据行（从起始行+1开始）
                table_data = []
                data_start_row = table_row + 1
                current_row = data_start_row
                
                while True:
                    # 检查第一列是否为空，为空则结束
                    first_cell = sheet.cell(row=current_row, column=table_col)
                    if first_cell.value is None:
                        break
                    
                    # 读取当前行数据
                    row_data = {}
                    for i, header in enumerate(headers):
                        value_cell = sheet.cell(row=current_row, column=table_col + i)
                        row_data[header] = value_cell.value
                    
                    table_data.append(row_data)
                    current_row += 1
                
                # 添加到tables_data
                tables_data[table_name] = table_data
            
            # 构建当前sheet的JSON对象
            sheet_data = {
                "text_data": text_data,
                "tables_data": tables_data
            }
            
            result.append(sheet_data)
        else:
            # 只有普通变量的情况：第一行为键，第二行为值，整个sheet生成一个文档
            # 处理text_data
            text_data = {}
            
            # text_data的键行和值行
            key_row = text_row
            value_row = text_row + 1
            current_col = text_col
            
            # 遍历所有列，直到遇到空单元格
            while True:
                key_cell = sheet.cell(row=key_row, column=current_col)
                value_cell = sheet.cell(row=value_row, column=current_col)
                
                # 检查是否为空单元格
                if key_cell.value is None:
                    break
                
                # 添加到text_data，键用{}包起来
                text_data[f"{{{key_cell.value}}}"] = value_cell.value
                current_col += 1
            
            # 构建当前sheet的JSON对象
            sheet_data = {
                "text_data": text_data,
                "tables_data": {}
            }
            
            result.append(sheet_data)
    
    return result
